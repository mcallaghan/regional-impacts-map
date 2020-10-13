import django, sys, os, re
from openpyxl import load_workbook
from django.core.exceptions import MultipleObjectsReturned, ObjectDoesNotExist
import pandas as pd

sys.path.append('/home/max/software/django-tmv/tmv_mcc-apsis/BasicBrowser')
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "BasicBrowser.settings")
django.setup()


from scoping.models import *
p = Project.objects.get(pk=178)
import bibtexparser
import json

os.chdir('/home/max/Documents/papers/regional-impacts-map/literature_identification')

# Load the IPCC .bib database
with open('IPCC/AR5 Chapter.bib') as bibtex_file:
    bib_database = bibtexparser.load(bibtex_file)




wb = load_workbook(filename = 'IPCC/Chapter 18 Masterfiles/TablesChapter18.xlsx')

u = User.objects.get(username="galm")

q, created = Query.objects.get_or_create(
        project = p,
        text = "MANUAL ADDED",
        title = "Table docs",
        database = "intern",
        creator = u
)

#q.doc_set.clear()

cat, created = Category.objects.get_or_create(
    project = p,
    name = "not included in table",
    description = "not included in table"
)

all_docs = Doc.objects.filter(docproject__project=p).exclude(category=cat)

def extract_refs(s):
    refs = None
    #for m in re.findall('\((.*)\)', s):
    p = '(?:\([^a-z]+\))*.*\((.*?[a-zA-Z]+.*?)\)(?:.*)(?:\([^a-z]+\))*'
    for m in re.findall(p, s):
        if re.match('.*[0-9]{4}',m):
            refs = re.split('([0-9]{4}[a-z]{0,1})[;,]*', m)
    if refs:
        return zip(refs[0::2], refs[1::2])
    else:
        print(s)
        return

def aupy_doc(au, py, title, all_docs):
    if " and " in au:
        auths = []
        for a in au.split(" and "):
            if title:
                docs = all_docs.filter(
                    docauthinst__AU__unaccent__regex=f'^{a}',
                    tslug=Doc.make_tslug(title),
                    PY=py
                )
            else:
                docs = all_docs.filter(
                    docauthinst__AU__unaccent__regex=f'^{a}',
                    PY=py
                )
            auths.append(set(docs.values_list('pk',flat=True)))
        dids = auths[0] & auths[1]
        docs = all_docs.filter(pk__in=dids)
        for d in docs:
            if d.docauthinst_set.distinct('AU').count() == 2:
                docs = all_docs.filter(pk=d.pk)
                break
    else:
        if title:
            docs = all_docs.filter(
                docauthinst__AU__unaccent__regex=f'^{au}',
                docauthinst__position__in=[0,1],
                tslug=Doc.make_tslug(title),
                PY=py
            )
        else:
            docs = all_docs.filter(
                docauthinst__AU__unaccent__regex=f'^{au}',
                docauthinst__position__in=[0,1],
                PY=py
            )
    if docs.distinct('id').count()==1:
        return docs.first()
    elif docs.distinct('id').count()==2 and len(set(docs.values_list('title',flat=True)))==1:
        docs = docs.filter(UT__UT__icontains="WOS:")
        if docs:
            return docs.first()

    elif docs.distinct('id').count()>1:
        print(f"\n{au} ({py}) matched the following documents")
        print(docs.distinct().values('docauthinst__AU','title','PY','UT__UT'))
    else:
        # Try once more with doi
        a1 = re.split(',| and ',au.strip())[0].replace(' ','')
        bib_key = f"{a1}{py}"
        e = [e for e in bib_database.entries if e['ID']==bib_key]
        if len(e)==1:
            if "doi" in e[0]:
                try:
                    doc = Doc.objects.get(wosarticle__di=e[0]["doi"])
                    return doc
                except MultipleObjectsReturned:
                    doc = Doc.objects.get(wosarticle__di=e[0]['doi'],UT__UT__contains="WOS")
                    return doc
                except:
                    pass

        return None
    
doc_matches = []

for ws in wb:
    system = None
    region = None
    for i, row in enumerate(ws.iter_rows()):
        if i == 0:
            pass
        elif i ==1:
            system = row[1].value
            syscat, created = Category.objects.get_or_create(
                project=p,
                name=system,
                description=f"system category: {system}",
                group="system",
            )
            print(f"\n\n#########\n{system}\n#########\n\n")
        else:
            if row[0].value:
                region = row[0].value
            refs = extract_refs(row[1].value)
            if refs:
                for au, py in refs:
                    doc = None
                    title = None
                    if len(py) > 4:
                        a1 = re.split('\W',au.strip())[0]
                        bib_key = f"{a1}{py}"
                        e = [e for e in bib_database.entries if e['ID']==bib_key][0]
                        title = e['title'].strip('{}')

                    py = py[:4]
                    au = re.split('\Wet',au)[0].strip(', ')

                    doc = aupy_doc(au, py, title, all_docs)
                    if not doc:
                        if " " in au:
                            au = au.replace(' ',', ')
                            doc = aupy_doc(au, py, title, all_docs)
                    if doc:
                        doc.query.add(q)
                        doc.category.add(syscat)
                    else:
                        if " " in au:
                            print(f"Couldn't match {au}, ({py}) or {au.replace(', ',' ')}, ({py})")
                        else:
                            print(f"Couldn't match {au}, ({py}). {title}")
                           
                    doc_id = None
                    if doc:
                        doc_id=doc.pk
                        
                    doc_matches.append({
                        "au": au,
                        "py": py,
                        "doc": doc_id,
                        "system": system,
                        "region": region
                    })
                    
            else:
                print(f"couldn't find refs in {row[1].value}")
                
doc_matches = pd.DataFrame.from_dict(doc_matches)
doc_matches.to_csv('IPCC_extraction.csv')

q.r_count = q.doc_set.count()
print(q.r_count)
q.save()
